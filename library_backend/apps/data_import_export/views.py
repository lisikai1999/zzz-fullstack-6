from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from django.contrib.auth.models import User
from openpyxl import Workbook, load_workbook
from io import BytesIO
from apps.books.models import Book, Category
from apps.users.permissions import IsAdmin, IsAdminOrLibrarian


@api_view(['GET'])
@permission_classes([IsAdminOrLibrarian])
def export_books(request):
    wb = Workbook()
    ws = wb.active
    ws.title = '图书列表'
    headers = ['ISBN', '书名', '作者', '出版社', '出版日期', '分类', '总册数', '可借册数', '馆藏位置', '状态']
    ws.append(headers)

    books = Book.objects.prefetch_related('categories').all()
    for book in books:
        categories = ', '.join([c.name for c in book.categories.all()])
        ws.append([
            book.isbn, book.title, book.author, book.publisher,
            book.publish_date.isoformat() if book.publish_date else '',
            categories, book.total_copies, book.available_copies,
            book.location, book.get_status_display()
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="books_export.xlsx"'
    return response


@api_view(['POST'])
@permission_classes([IsAdminOrLibrarian])
@parser_classes([MultiPartParser])
def import_books(request):
    file = request.FILES.get('file')
    if not file:
        return Response({'detail': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        wb = load_workbook(file)
        ws = wb.active
    except Exception:
        return Response({'detail': '无法解析文件'}, status=status.HTTP_400_BAD_REQUEST)

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        isbn, title, author, publisher, publish_date, categories_str, total, available, location, _ = (
            list(row) + [None] * 10)[:10]

        if not isbn or not title:
            errors.append(f'第{i}行: ISBN和书名不能为空')
            continue

        if Book.objects.filter(isbn=str(isbn)).exists():
            errors.append(f'第{i}行: ISBN {isbn} 已存在')
            continue

        book = Book.objects.create(
            isbn=str(isbn).strip(),
            title=str(title).strip(),
            author=str(author).strip() if author else '',
            publisher=str(publisher).strip() if publisher else '',
            total_copies=int(total) if total else 1,
            available_copies=int(available) if available else 1,
            location=str(location).strip() if location else '',
        )

        if categories_str:
            for cat_name in str(categories_str).split(','):
                cat_name = cat_name.strip()
                if cat_name:
                    cat, _ = Category.objects.get_or_create(name=cat_name)
                    book.categories.add(cat)
        created += 1

    return Response({
        'created': created,
        'errors': errors,
        'total_rows': len(rows),
    })


@api_view(['GET'])
@permission_classes([IsAdmin])
def export_readers(request):
    wb = Workbook()
    ws = wb.active
    ws.title = '读者列表'
    headers = ['用户名', '姓名', '邮箱', '电话', '会员类型', '借书证号', '地址']
    ws.append(headers)

    users = User.objects.filter(profile__role='reader').select_related('profile')
    for user in users:
        ws.append([
            user.username,
            f"{user.last_name}{user.first_name}",
            user.email,
            user.profile.phone,
            user.profile.get_membership_type_display(),
            user.profile.id_card,
            user.profile.address,
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="readers_export.xlsx"'
    return response


@api_view(['POST'])
@permission_classes([IsAdmin])
@parser_classes([MultiPartParser])
def import_readers(request):
    file = request.FILES.get('file')
    if not file:
        return Response({'detail': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        wb = load_workbook(file)
        ws = wb.active
    except Exception:
        return Response({'detail': '无法解析文件'}, status=status.HTTP_400_BAD_REQUEST)

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        username, name, email, phone, membership, id_card, address = (list(row) + [None] * 7)[:7]

        if not username:
            errors.append(f'第{i}行: 用户名不能为空')
            continue

        if User.objects.filter(username=str(username).strip()).exists():
            errors.append(f'第{i}行: 用户名 {username} 已存在')
            continue

        last_name = str(name)[0] if name and len(str(name)) > 0 else ''
        first_name = str(name)[1:] if name and len(str(name)) > 1 else ''

        user = User.objects.create_user(
            username=str(username).strip(),
            password='123456',
            email=str(email).strip() if email else '',
            last_name=last_name,
            first_name=first_name,
        )
        profile = user.profile
        profile.phone = str(phone).strip() if phone else ''
        profile.id_card = str(id_card).strip() if id_card else ''
        profile.address = str(address).strip() if address else ''

        membership_map = {'标准': 'standard', '高级': 'premium', '学生': 'student'}
        if membership and str(membership) in membership_map:
            profile.membership_type = membership_map[str(membership)]

        profile.save()
        created += 1

    return Response({
        'created': created,
        'errors': errors,
        'total_rows': len(rows),
    })


@api_view(['GET'])
@permission_classes([IsAdminOrLibrarian])
def template_books(request):
    wb = Workbook()
    ws = wb.active
    ws.title = '图书导入模板'
    headers = ['ISBN', '书名', '作者', '出版社', '出版日期', '分类(逗号分隔)', '总册数', '可借册数', '馆藏位置']
    ws.append(headers)
    ws.append(['9787111111111', '示例图书', '张三', '人民出版社', '2024-01-01', '文学,小说', '3', '3', 'A-1-01'])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="books_template.xlsx"'
    return response


@api_view(['GET'])
@permission_classes([IsAdmin])
def template_readers(request):
    wb = Workbook()
    ws = wb.active
    ws.title = '读者导入模板'
    headers = ['用户名', '姓名', '邮箱', '电话', '会员类型(标准/高级/学生)', '借书证号', '地址']
    ws.append(headers)
    ws.append(['zhangsan', '张三', 'zhangsan@example.com', '13800138000', '标准', 'LIB001', '北京市海淀区'])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="readers_template.xlsx"'
    return response
